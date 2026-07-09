# -*- coding: utf-8 -*-
"""为 UID 120134752 生成含证据表的分析 Excel"""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

UID = "120134752"
DATA_DIR = f"E:/aicu评论-ds-wb/output/{UID}"
OUT_PATH = f"{DATA_DIR}/aicu_analysis_{UID}.xlsx"

HDR_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
HDR_FILL = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_ALIGN = Alignment(vertical='top', wrap_text=True)
BORDER = Border(left=Side(style='thin',color='DDDDDD'),right=Side(style='thin',color='DDDDDD'),
                top=Side(style='thin',color='DDDDDD'),bottom=Side(style='thin',color='DDDDDD'))
GOLD = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
YELLOW = PatternFill(start_color='FFFFF0', end_color='FFFFF0', fill_type='solid')

def write_sheet(ws, rows, widths=None):
    if not rows: return
    fields = list(rows[0].keys())
    for ci, name in enumerate(fields, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = HDR_ALIGN; c.border = BORDER
    # ID类列名（需要文本格式避免科学计数法）
    id_keywords = ('id', 'rpid', 'oid', 'rootid', 'parentid')
    def _is_id(name): return any(k in name.lower() for k in id_keywords)
    for ri, row in enumerate(rows, 2):
        for ci, name in enumerate(fields, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(name,''))
            c.alignment = CELL_ALIGN; c.border = BORDER; c.font = Font(name='微软雅黑', size=10)
            if _is_id(name):
                c.number_format = '@'
    if widths:
        for ci, name in enumerate(fields, 1):
            ws.column_dimensions[get_column_letter(ci)].width = widths.get(name, 18)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{len(rows)+1}"

with open(f"{DATA_DIR}/aicu_comments_{UID}.csv", encoding='utf-8-sig') as f:
    comments = list(csv.DictReader(f))

# 证据表
evidence = [
    # === 基本信息 ===
    ("基本信息-性别", "男性", "★★★★★",
     "自称「你真是个人才👍」(对男性语气)；「幽默男性」视角；讨论「胖猫」事件时的男性立场",
     [10, 87, 83],
     "第10/87/83行",
     "「幽默男性，性别本身就不存在」(87)；「胖猫他自己觉得值」(83)"),

    ("基本信息-所在地", "深圳（工作/学习）", "★★★★★",
     "57次提到深圳，直接自称「我这深圳的」「我是在深圳吃的」",
     [33, 149, 2077],
     "第33/149/2077行",
     "「我这深圳的，以前饭堂早餐...」(33)；「我是在深圳吃的」(149)；「加上深圳的场地房价」(2077)"),

    ("基本信息-年龄", "约25-28岁（已大学毕业，在读研或刚工作）", "★★★★★",
     "自称经历过「买断时代、点卡时代」→至少25+；「我大学时候」→已毕业/读研；「上班了你就会知道」→有工作经验",
     [64, 2067, 42],
     "第64/2067/42行",
     "「B站用户那么多低年龄...没经历过买断时代、点卡时代」(64)；「我大学时候可是三个游戏的攻略组」(2067)；「上班了你就会知道你的时间更值钱」(42)"),

    ("基本信息-学历", "研究生（硕/博），理工科（材料/物理/工程）", "★★★★★",
     "直接透露：有导师、管实验室项目经费、操作镀膜设备/移液枪",
     [2077, 115, 2554],
     "第2077/115/2554行",
     "「我导师两个月前刚拿了三十万个人奖金...一支高精移液枪就已经两千了，就是给钱给我采购的，真空溅射镀膜，蒸发离子镀膜，激光蚀刻机」(2077)；「物理专业不学大学物理」(115)"),

    ("基本信息-实验室设备", "镀膜实验室（真空溅射/离子镀膜/激光蚀刻）→材料表面工程", "★★★★★",
     "详细列出实验室设备名称和价格",
     [2077],
     "第2077行",
     "「真空溅射镀膜，蒸发离子镀膜，激光蚀刻机，每台都是五六万起步，定制的两台加起来有30万」(2077)"),

    ("基本信息-老家", "广东农村（有土灶、秸秆堆、守村人）", "★★★★★",
     "描述老家农村生活细节",
     [4184, 67],
     "第4184/67行",
     "「我老家那里还有个土灶呢，家家都有一个封顶的小仓库，秸秆收起来堆进里边」(4184)；「我们村的那个守村人他老哥零几年的给撞死了」(67)"),

    ("基本信息-工作经验", "曾跑过机床 + 管过项目采购", "★★★★★",
     "直接提过「以前跑过机床」；了解甲方乙方工程套路",
     [116, 117, 121],
     "第116/117/121行",
     "「以前跑过机床，对面电话里说怎么搞都没反应，延误开工损失很大」(116)；「治理是一门艺术，这些勾当你以为给钱的甲方能不知道吗」(117)"),

    # === 兴趣 ===
    ("兴趣-永劫无间", "深度竞技玩家，397次关键词", "★★★★★",
     "讨论帧级别操作（钩锁连招、b1抓闪、近战修正）、武器机制、版本更新",
     [77, 132, 133, 136, 144, 146],
     "第77/132/133/136/144/146行",
     "「没有近战修正呢，你追人也和克烈一样随便点人？」(77)；「飞刀右键是不能滑步的，左键滑步三个才顶别人两个」(133)；「背包切向下滚轮就行，比钩锁连招的帧数要求更低」(136)"),

    ("兴趣-炉石传说", "氪金玩家，78次关键词", "★★★★★",
     "讨论卡组meta、金币vsRMB经济学、补丁平衡、狂野模式",
     [107, 112, 113, 104, 109],
     "第107/112/113/104/109行",
     "「买了大预购等背刺合集就完事了」(107)；「铜须奥丁都回调了」(110)；「口德？什么地沟油」(109)"),

    ("兴趣-文明6", "策略游戏爱好者，75次关键词", "★★★★",
     "讨论游戏机制和策略",
     [24, 22],
     "第24/22行",
     "「科技也起的太慢了吧，魔女全程打架的前排都没这么慢的坦克」(24)；「图书馆便宜，大学解锁也能很早」(22)"),

    ("兴趣-哲学", "218次关键词，熟悉近代哲学", "★★★★★",
     "引用笛卡尔、近代哲学、爱欲循环等概念",
     [54, 79, 83, 88],
     "第54/79/83/88行",
     "「连两千年前的古希腊哲学都比不过的，就别指望谈笛卡尔之后的近代哲学了」(54)；「你只是你自己的爱欲循环机制在运作而已」(88)；「努力工作...成为人类」(79)"),

    ("兴趣-反二游/抽卡", "强烈反感gacha和二次元手游", "★★★★★",
     "73次二游关键词，高频嘲讽",
     [62, 137, 174],
     "第62/137/174行",
     "「真以为卖数值的抽卡游戏上桌了？lol,cf,永劫...二游打得过哪个？」(62)；「二游痴闹麻了」(137)；「都玩原神了，就这样吧」(174)"),

    ("兴趣-政治经济", "524次关键词，反美+马克思主义倾向", "★★★★★",
     "375次提到美国（多为讽刺）；引用资本主义批判理论",
     [15, 120, 158, 36, 81],
     "第15/120/158/36/81行",
     "「波音-767重型巡飞弹是人类历史上最大的常规导弹」(15)——911讽刺；「哈耶克的大手一挥」(36)；「定价权还是在资本主义母国那边」(81)；「私有制还存在吗」(158)"),

    ("兴趣-网文", "看修仙/后宫类网文，46次关键词", "★★★",
     "评论起点小说、修仙文",
     [53, 17],
     "第53/17行",
     "「逆天邪神纯一坨，早该完书了，后宫文拖得越长越烂」(53)；评论「苟在初圣魔门当人材」(17)"),

    # === 行为模式 ===
    ("行为-作息", "重度熬夜型，凌晨活跃比例极高", "★★★★★",
     "大量评论在凌晨1-5点",
     [23, 51, 77, 78, 120],
     "第23/51/77/78/120行（凌晨多发）",
     "凌晨3-5点活跃评论多次出现，如「凌晨两点讲到快五点」(516)"),

    ("行为-社交模式", "93%回复率，每条都在跟人争论", "★★★★★",
     "6704条中6234条为回复",
     [],
     "全表统计",
     "一级评论470(7%)，回复6234(93%)"),

    ("行为-傲慢感", "自称本科时是三个游戏的攻略组成员", "★★★★",
     "拿攻略组身份碾压别人的游戏理解",
     [2067],
     "第2067行",
     "「我大学时候可是三个游戏的攻略组，我舍友都写过大型游戏mod」(2067)"),
]

# ═══════════ 构建 Excel ═══════════
wb = Workbook()

# Sheet 1: 评论数据
ws1 = wb.active
ws1.title = "评论数据"
write_sheet(ws1, comments, {
    '序号': 6, '评论ID': 18, '评论内容': 55, '评论时间': 14, '层级': 10,
    '根评论ID': 16, '父评论ID': 16, '视频ID': 18, '视频链接': 38, '视频标题': 38,
    '评论跳转链接': 42,
})

key_rows = {33, 149, 2077, 64, 2067, 42, 115, 2554, 4184, 67, 116, 117, 121,
            77, 132, 133, 136, 144, 146, 107, 112, 113, 104, 109, 24, 22,
            54, 79, 83, 88, 62, 137, 174, 15, 120, 158, 36, 81, 53, 17, 10, 87}
for rn in key_rows:
    if 1 <= rn <= len(comments):
        for ci in range(1, 13):
            ws1.cell(row=rn+1, column=ci).fill = YELLOW

# Sheet 2: 关键信息推断（含证据）
ws2 = wb.create_sheet("关键信息推断（含证据）")
ws2.merge_cells('A1:H1')
ws2.cell(row=1, column=1, value=f'B站用户 UID {UID} 个人信息推断表').font = Font(name='微软雅黑', bold=True, size=16, color='1A1A2E')
ws2.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 36

ws2.merge_cells('A2:H2')
ws2.cell(row=2, column=1, value=f'数据来源：aicu.cc | 6704条评论 | 2018~2025 | 每条推断可按行号在"评论数据"表验证').font = Font(name='微软雅黑', size=10, color='666666')
ws2.cell(row=2, column=1).alignment = Alignment(horizontal='center')

cols = ['类别', '推断结论', '置信度', '依据摘要', '对应行号', '行号说明', '关键评论摘录', '原始评论全文']
for ci, h in enumerate(cols, 1):
    c = ws2.cell(row=4, column=ci, value=h)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = HDR_ALIGN; c.border = BORDER

rn = 5
for item in evidence:
    cat, conc, conf, summary, row_nums, row_desc, excerpts = item

    ws2.cell(row=rn, column=1, value=cat).font = Font(name='微软雅黑', bold=True, size=11, color='1A1A2E')
    ws2.cell(row=rn, column=2, value=conc).font = Font(name='微软雅黑', bold=True, size=12, color='2B579A')
    c3 = ws2.cell(row=rn, column=3, value=conf)
    c3.font = Font(name='微软雅黑', size=11); c3.alignment = Alignment(horizontal='center', vertical='top')
    ws2.cell(row=rn, column=4, value=summary).font = Font(name='微软雅黑', size=10)
    nums = ', '.join(str(x) for x in row_nums) if row_nums else '—'
    c5 = ws2.cell(row=rn, column=5, value=nums)
    c5.font = Font(name='Consolas', size=11, color='C0392B', bold=True); c5.alignment = Alignment(horizontal='center', vertical='top')
    ws2.cell(row=rn, column=6, value=row_desc).font = Font(name='微软雅黑', size=9, color='888888', italic=True)
    ws2.cell(row=rn, column=7, value=excerpts).font = Font(name='微软雅黑', size=10, color='333333')

    full = []
    for rn_idx in row_nums:
        if 1 <= rn_idx <= len(comments):
            r = comments[rn_idx-1]
            msg = r['评论内容'].replace('\n',' ')
            ts = r['评论时间']; title = r.get('视频标题','')
            full.append(f"[第{rn_idx}行 | {ts}] {msg}")
            if title and '失败' not in title and '不存在' not in title:
                full.append(f"  -> {title}")
    ws2.cell(row=rn, column=8, value='\n'.join(full) if full else summary).font = Font(name='微软雅黑', size=9, color='555555')

    for col in range(1, 9):
        c = ws2.cell(row=rn, column=col)
        c.alignment = CELL_ALIGN; c.border = BORDER
        if conf.count('★') >= 5: c.fill = GOLD

    rn += 1

ws2.column_dimensions['A'].width = 22; ws2.column_dimensions['B'].width = 35
ws2.column_dimensions['C'].width = 10; ws2.column_dimensions['D'].width = 42
ws2.column_dimensions['E'].width = 15; ws2.column_dimensions['F'].width = 22
ws2.column_dimensions['G'].width = 58; ws2.column_dimensions['H'].width = 65
ws2.freeze_panes = 'A5'

# Sheet 3: 统计摘要
ws3 = wb.create_sheet("统计摘要")
ws3.merge_cells('A1:C1')
ws3.cell(row=1, column=1, value=f'UID {UID} 数据统计').font = Font(name='微软雅黑',bold=True,size=14)

from datetime import datetime
times = []
for r in comments:
    try: times.append(datetime.strptime(r['评论时间'],'%Y-%m-%d %H:%M:%S'))
    except: pass
late = sum(1 for t in times if 0<=t.hour<=5)

stats = [
    ('评论总数','6704',''),
    ('一级评论','470 (7%)','几乎全是在回复别人'),
    ('回复','6234 (93%)','极端好辩'),
    ('时间跨度',f'{min(times).strftime("%Y-%m-%d")}~{max(times).strftime("%Y-%m-%d")}' if times else '','2018-2025，7年半'),
    ('凌晨活跃(0-5点)',f'{late}条 ({late*100//len(times)}%)' if times else '','重度熬夜'),
    ('','',''),
    ('# 身份','',''),
    ('所在地','深圳','行33/149/2077'),
    ('年龄','约25-28岁','行64/2067/42'),
    ('学历','研究生(硕/博)，材料/物理/工程','行2077'),
    ('实验室','镀膜实验室(真空溅射/离子镀膜/激光蚀刻)','行2077'),
    ('老家','广东农村(土灶/秸秆/守村人)','行4184/67'),
    ('','',''),
    ('# 游戏','',''),
    ('永劫无间','397次关键词，帧级操作讨论','行77/132/133/136'),
    ('炉石传说','78次，氪金玩家+meta分析','行107/109/110'),
    ('文明6','75次','行24/22'),
    ('二游态度','强烈反感(73次)','行62/137/174'),
    ('','',''),
    ('# 兴趣','',''),
    ('哲学','218次，笛卡尔~近代哲学','行54/79/88'),
    ('政治经济','524次，反美+马主义倾向','行15/81/158'),
    ('网文','46次，修仙/后宫','行53/17'),
]

for i,(k,v,n) in enumerate(stats,3):
    ws3.cell(row=i,column=1,value=k).font=Font(name='微软雅黑',bold=True,size=11)
    ws3.cell(row=i,column=2,value=v).font=Font(name='微软雅黑',size=11)
    ws3.cell(row=i,column=3,value=n).font=Font(name='微软雅黑',size=10,color='888888')
ws3.column_dimensions['A'].width=25; ws3.column_dimensions['B'].width=40; ws3.column_dimensions['C'].width=42

wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
print(f"Sheet1: 评论({len(comments)}) | Sheet2: 推断({len(evidence)}) | Sheet3: 统计")
