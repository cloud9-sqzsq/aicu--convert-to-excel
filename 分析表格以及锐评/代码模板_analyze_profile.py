"""
B站用户评论个人画像分析 — 可复用代码模板
===========================================

使用方法：
  1. 填写 ANALYSIS 字典（分析结论 + 锐评文本）
  2. 更新 HOLIDAYS 字典（当前年份的放假安排）
  3. 运行 python 本文件.py
  4. 在 output/ 目录获取生成的 Excel

依赖：openpyxl（需安装到 libs/ 目录）
输入：output/ 目录下的 B站评论 CSV
输出：output/个人信息分析_锐评版_{timestamp}.xlsx
"""

import csv
import json
import os
import sys
from datetime import datetime
from collections import Counter, defaultdict

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "libs"))


# =============================================================================
# 第一部分：分析内容配置（需要手动编写）
# =============================================================================

ANALYSIS = {
    "基本信息": [
        {
            "维度": "性别",
            "推断结论": "待填写",
            "置信度": "高",  # 高 / 中 / 低 / 极高
            "锐评": "待填写——讽刺风格的评论，100-300字",
            "依据行号": [],  # CSV 中对应的行号列表 [12, 34, 56]
            "原文摘要": "待填写——引用评论原文的关键词句",
        },
        # ... 添加更多维度：年龄段、所在地、手机型号、身高体重、家庭关系
    ],
    "职业与学业": [
        {
            "维度": "职业身份",
            "推断结论": "待填写",
            "置信度": "极高",
            "锐评": "待填写",
            "依据行号": [],
            "原文摘要": "待填写",
        },
        # ... 学习状态、学业方向、知识面
    ],
    "兴趣爱好": [
        {
            "维度": "兴趣1",
            "推断结论": "待填写",
            "置信度": "高",
            "锐评": "待填写",
            "依据行号": [],
            "原文摘要": "待填写",
        },
        # ... 每个兴趣单独一项
    ],
    "性格特征": [
        {
            "维度": "网络攻击性",
            "推断结论": "待填写",
            "置信度": "极高",
            "锐评": "待填写",
            "依据行号": [],
            "原文摘要": "待填写",
        },
        # ... 理性程度、自我认知、价值观
    ],
    "总体画像（锐评总结）": [
        {
            "维度": "一句话总结",
            "推断结论": "待填写——一段话概括这个人的完整形象",
            "置信度": "较真你就输了",
            "锐评": "待填写——最终毒舌总结",
            "依据行号": [1, 999],  # 开始行到结束行
            "原文摘要": "从头到尾，每一条都是ta自己写的",
        },
    ],
}


# =============================================================================
# 第二部分：节假日数据配置（每年更新）
# =============================================================================

HOLIDAYS = {
    # 格式: 'YYYY-MM-DD': '节日名' 或 '调休上班'
    # ==== 2025 年 ====
    "2025-01-01": "元旦",
    "2025-01-28": "春节", "2025-01-29": "春节", "2025-01-30": "春节",
    "2025-01-31": "春节",
    "2025-02-01": "春节", "2025-02-02": "春节", "2025-02-03": "春节",
    "2025-02-04": "春节",
    "2025-01-26": "调休上班",
    "2025-02-08": "调休上班",
    "2025-04-04": "清明", "2025-04-05": "清明", "2025-04-06": "清明",
    "2025-05-01": "劳动节", "2025-05-02": "劳动节", "2025-05-03": "劳动节",
    "2025-05-04": "劳动节", "2025-05-05": "劳动节",
    "2025-04-27": "调休上班",
    "2025-05-31": "端午", "2025-06-01": "端午", "2025-06-02": "端午",
    "2025-10-01": "国庆中秋", "2025-10-02": "国庆中秋",
    "2025-10-03": "国庆中秋", "2025-10-04": "国庆中秋",
    "2025-10-05": "国庆中秋", "2025-10-06": "国庆中秋",
    "2025-10-07": "国庆中秋", "2025-10-08": "国庆中秋",
    "2025-09-28": "调休上班", "2025-10-11": "调休上班",
    # ==== 在这里添加更多年份 ====
}

WEEKDAY_NAMES = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]


# =============================================================================
# 第三部分：核心工具函数（不需要修改）
# =============================================================================

def load_csv_rows():
    """从 output/ 目录读取最新的评论 CSV"""
    csv_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
    csv_files = [f for f in os.listdir(csv_dir) if f.endswith(".csv")]
    if not csv_files:
        print("错误：未找到CSV文件")
        return [], None
    csv_path = os.path.join(csv_dir, csv_files[0])
    rows = []
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows, csv_files[0]


def _is_school_break(dt):
    """判断寒暑假：寒假 1月15日-2月28日，暑假 7-8月"""
    m, d = dt.month, dt.day
    if m in (7, 8):
        return True
    if m == 1 and d >= 15:
        return True
    if m == 2 and d <= 28:
        return True
    return False


def _classify_day_full(dt):
    """完整日期分类：法定假日 > 调休 > 寒暑假 > 周末 > 工作日
    返回: (详细分类, 大类, 是否真工作日, 假日名称或None)
    """
    date_str = dt.strftime("%Y-%m-%d")
    wd = dt.weekday()
    htag = HOLIDAYS.get(date_str, None)

    if htag and "调休" not in htag:
        return (f"法定假日({htag})", "法定假日", False, htag)
    elif htag and "调休" in htag:
        return ("调休上班日(周末上班)", "工作日（含调休）", True, None)
    elif _is_school_break(dt):
        return ("寒暑假", "寒暑假", False, None)
    elif wd in (5, 6):
        return ("周末", "周末", False, None)
    else:
        return ("工作日", "工作日（含调休）", True, None)


# =============================================================================
# 第四部分：Excel 生成函数（按需修改布局）
# =============================================================================

def generate_time_analysis_sheet(wb, rows):
    """生成"评论时间分析"分表"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("评论时间分析")

    # 样式定义
    title_font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="B22222", end_color="B22222", fill_type="solid")
    section_font = Font(name="微软雅黑", size=13, bold=True, color="8B0000")
    section_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
    roast_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
    normal_font = Font(name="微软雅黑", size=10)
    roast_font = Font(name="微软雅黑", size=10, color="8B0000")
    bold_font = Font(name="微软雅黑", size=10, bold=True)
    number_font = Font(name="微软雅黑", size=10, bold=True)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="top")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 解析时间戳，对每条评论做分类
    results = []
    for i, r in enumerate(rows, 1):
        ts = int(r.get("时间戳", 0))
        if not ts:
            continue
        dt = datetime.fromtimestamp(ts)
        detail_cat, group, is_workday, holiday_name = _classify_day_full(dt)
        h = dt.hour
        if 0 <= h < 8:
            time_slot = "深夜/凌晨(0-8)"
        elif 8 <= h < 12:
            time_slot = "上午(8-12)"
        elif 12 <= h < 14:
            time_slot = "午休(12-14)"
        elif 14 <= h < 17:
            time_slot = "下午上课/工作(14-17)"
        elif 17 <= h < 19:
            time_slot = "傍晚(17-19)"
        elif 19 <= h < 22:
            time_slot = "晚间(19-22)"
        else:
            time_slot = "深夜(22-24)"

        results.append({
            "row": i, "dt": dt, "date": dt.strftime("%Y-%m-%d"),
            "weekday": dt.weekday(),
            "weekday_name": WEEKDAY_NAMES[dt.weekday()],
            "hour": h, "detail_cat": detail_cat, "group": group,
            "is_workday": is_workday, "time_slot": time_slot,
            "holiday_name": holiday_name,
            "in_class": is_workday and ((9 <= h < 12) or (14 <= h < 17)),
        })

    total = len(results)
    if total == 0:
        ws.cell(row=1, column=1, value="无时间数据").font = normal_font
        return

    # ---- 分类汇总 ----
    group_cnt = Counter(r["group"] for r in results)
    detail_cnt = Counter(r["detail_cat"] for r in results)
    hour_type_counts = Counter(r["time_slot"] for r in results)
    workday_hour_cnt = Counter(r["time_slot"] for r in results if r["is_workday"])

    real_workday_total = sum(1 for r in results if r["is_workday"])
    workday_class = sum(1 for r in results if r["is_workday"] and r["in_class"])
    workday_off = real_workday_total - workday_class
    tiaoxiu_list = [r for r in results if "调休" in r["detail_cat"]]
    holiday_list = [r for r in results if "法定假日" in r["detail_cat"]]
    summer_cnt = sum(1 for r in results if r["dt"].month in (7, 8))
    winter_cnt = sum(1 for r in results if r["dt"].month in (1, 2))

    row_idx = 1

    # ---- 大标题 ----
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
    ws.cell(row=row_idx, column=1, value="评论时间分析（锐评版）").font = title_font
    ws.cell(row=row_idx, column=1).fill = title_fill
    ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row_idx].height = 40
    row_idx += 2

    # ---- 第一节：日期类型总览 ----
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
    ws.cell(row=row_idx, column=1, value="一、评论日期类型总览").font = section_font
    ws.cell(row=row_idx, column=1).fill = section_fill
    row_idx += 1

    for h_text in ["日期类型", "评论数", "占比", "锐评"]:
        ci = ["日期类型", "评论数", "占比", "锐评"].index(h_text) + 1
        ws.cell(row=row_idx, column=ci, value=h_text).font = header_font
        ws.cell(row=row_idx, column=ci).fill = header_fill
        ws.cell(row=row_idx, column=ci).alignment = center_align
        ws.cell(row=row_idx, column=ci).border = thin_border
    row_idx += 1

    for dtype in ["寒暑假", "工作日（含调休）", "周末", "法定假日"]:
        cnt = group_cnt.get(dtype, 0)
        pct = cnt / total * 100 if total else 0
        ws.cell(row=row_idx, column=1, value=dtype).font = bold_font
        ws.cell(row=row_idx, column=2, value=cnt).font = number_font
        ws.cell(row=row_idx, column=3, value=f"{pct:.1f}%").font = normal_font
        ws.cell(row=row_idx, column=4, value=f"共计{cnt}条").font = roast_font
        ws.cell(row=row_idx, column=4).fill = roast_fill
        for ci in range(1, 5):
            ws.cell(row=row_idx, column=ci).border = thin_border
        row_idx += 1

    row_idx += 1

    # ---- 调休说明 ----
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
    ws.cell(row=row_idx, column=1, value="调休修正明细").font = Font(name="微软雅黑", size=11, bold=True, color="8B0000")
    row_idx += 1

    tiaoxiu_headers = ["类型", "日期", "星期", "原分类", "修正后", "评论数"]
    for ci, h in enumerate(tiaoxiu_headers, 1):
        ws.cell(row=row_idx, column=ci, value=h).font = header_font
        ws.cell(row=row_idx, column=ci).fill = header_fill
        ws.cell(row=row_idx, column=ci).alignment = center_align
        ws.cell(row=row_idx, column=ci).border = thin_border
    row_idx += 1

    if tiaoxiu_list:
        tiaoxiu_by_date = defaultdict(list)
        for r in tiaoxiu_list:
            tiaoxiu_by_date[r["date"]].append(r)
        for date_str, items in sorted(tiaoxiu_by_date.items()):
            item = items[0]
            ws.cell(row=row_idx, column=1, value="调休上班日").font = bold_font
            ws.cell(row=row_idx, column=2, value=date_str).font = normal_font
            ws.cell(row=row_idx, column=3, value=item["weekday_name"]).font = normal_font
            ws.cell(row=row_idx, column=4, value="→周末").font = Font(name="微软雅黑", size=10, color="999999")
            ws.cell(row=row_idx, column=5, value="→工作日").font = Font(name="微软雅黑", size=10, color="B22222", bold=True)
            ws.cell(row=row_idx, column=6, value=len(items)).font = number_font
            for ci in range(1, 7):
                ws.cell(row=row_idx, column=ci).border = thin_border
                ws.cell(row=row_idx, column=ci).alignment = center_align
            row_idx += 1

    # ---- 列宽 ----
    for i, w in enumerate([22, 14, 18, 14, 14, 64], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# （篇幅限制，完整版参见项目中的 analyze_profile.py）


def generate_excel(rows, csv_name):
    """生成完整 Excel 报告"""
    try:
        import openpyxl
    except ImportError:
        import subprocess
        subprocess.check_call([
            sys.executable, "-m", "pip", "install",
            f"--target={os.path.join(os.path.dirname(os.path.abspath(__file__)), 'libs')}",
            "openpyxl",
        ])
        import openpyxl

    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ==== Sheet 1: 个人信息分析 ====
    ws = wb.active
    ws.title = "个人信息分析"

    title_font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="B22222", end_color="B22222", fill_type="solid")
    category_font = Font(name="微软雅黑", size=13, bold=True, color="8B0000")
    category_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
    roast_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
    normal_font = Font(name="微软雅黑", size=10)
    roast_font = Font(name="微软雅黑", size=10, color="8B0000")
    bold_font = Font(name="微软雅黑", size=10, bold=True)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="top")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    row_idx = 1

    # 大标题
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
    cell = ws.cell(row=row_idx, column=1,
                   value="B站用户评论个人信息分析报告（锐评版）")
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row_idx].height = 40
    row_idx += 1

    # 副标题
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
    cell = ws.cell(row=row_idx, column=1,
                   value=f"数据来源: {csv_name} | 共 {len(rows)} 条评论 | "
                         f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
                         f"免责声明：本分析仅供娱乐")
    cell.font = Font(name="微软雅黑", size=9, color="999999")
    cell.alignment = Alignment(horizontal="center")
    row_idx += 2

    # 活跃时间概况（此处省略详细实现，参见完整版）
    # ...

    # 逐类别输出分析
    for category, items in ANALYSIS.items():
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        cell = ws.cell(row=row_idx, column=1, value=category)
        cell.font = category_font
        cell.fill = category_fill
        ws.row_dimensions[row_idx].height = 30
        row_idx += 1

        sub_headers = ["维度", "推断结论", "置信度", "锐评", "依据行号", "原文摘要"]
        for ci, h in enumerate(sub_headers, 1):
            cell = ws.cell(row=row_idx, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row_idx += 1

        for item in items:
            row_vals = [
                item["维度"], item["推断结论"], item["置信度"],
                item["锐评"],
                "、".join(str(r) for r in item["依据行号"]),
                item["原文摘要"],
            ]
            for ci, val in enumerate(row_vals, 1):
                cell = ws.cell(row=row_idx, column=ci, value=val)
                cell.font = roast_font if ci == 4 else normal_font
                cell.alignment = wrap_align if ci >= 4 else center_align
                cell.border = thin_border
                if ci == 4:
                    cell.fill = roast_fill
            ws.row_dimensions[row_idx].height = max(60, len(str(item.get("锐评", ""))) // 50 * 18 + 30)
            row_idx += 1
        row_idx += 1

    # 列宽
    for i, w in enumerate([14, 30, 8, 68, 20, 64], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ==== Sheet 2: 原文索引 ====
    ws2 = wb.create_sheet("关键评论原文索引")
    # ... （详见完整版 analyze_profile.py）

    # ==== 保存 ====
    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"个人信息分析_锐评版_{timestamp}.xlsx")

    generate_time_analysis_sheet(wb, rows)  # Sheet 3
    wb.save(filepath)
    print(f"分析报告已保存: {filepath}")
    return filepath


# =============================================================================
# 第五部分：主入口
# =============================================================================

def main():
    print("正在读取评论数据...")
    rows, csv_name = load_csv_rows()
    if not rows:
        print("错误：无评论数据")
        return
    print(f"读取到 {len(rows)} 条评论，开始锐评分析...")
    generate_excel(rows, csv_name)
    print("锐评完成！")


if __name__ == "__main__":
    main()
