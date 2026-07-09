# -*- coding: utf-8 -*-
"""生成完整多表 Excel：评论 + 视频弹幕 + 直播弹幕 + 粉丝牌"""

import csv, json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

UID = "1703916229"
OUT_DIR = "E:/aicu评论-ds-wb/output"
CACHE_DIR = OUT_DIR

# ── 公共样式 ──
HEADER_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_ALIGN = Alignment(vertical='top', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD'),
)
HIGHLIGHT_FILL = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')

def write_sheet(ws, rows, col_widths=None):
    """通用写入 Sheet"""
    if not rows:
        ws.cell(row=1, column=1, value="(暂无数据)").font = Font(name='微软雅黑', size=12, color='999999')
        return
    fieldnames = list(rows[0].keys())
    # 表头
    for ci, name in enumerate(fieldnames, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = HEADER_ALIGN; c.border = THIN_BORDER
    # ID类列名（需要文本格式避免科学计数法）
    id_col_names = {'ID', '(oid)', 'rpid', 'oid', 'rootid', 'parentid'}
    def _is_id_col(name: str) -> bool:
        return any(k in name.lower() for k in id_col_names)
    # 数据
    for ri, row in enumerate(rows, 2):
        for ci, name in enumerate(fieldnames, 1):
            val = row.get(name, '')
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = CELL_ALIGN; c.border = THIN_BORDER; c.font = Font(name='微软雅黑', size=10)
            if _is_id_col(name):
                c.number_format = '@'
    # 列宽
    if col_widths:
        for ci, name in enumerate(fieldnames, 1):
            w = col_widths.get(name, 18)
            ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}{len(rows)+1}"


def safe_json_load(path, default=None):
    if os.path.exists(path):
        with open(path, encoding='utf-8') as f:
            return json.load(f)
    return default or []


# ═══════════════════════════════════════
# 1. 评论数据（从 CSV 读取）
# ═══════════════════════════════════════
def load_comments():
    csv_path = f"{OUT_DIR}/aicu_comments_{UID}.csv"
    if not os.path.exists(csv_path):
        return []
    with open(csv_path, encoding='utf-8-sig') as f:
        return list(csv.DictReader(f))


# ═══════════════════════════════════════
# 2. 粉丝牌
# ═══════════════════════════════════════
def load_medals():
    json_path = f"{OUT_DIR}/aicu_medals_raw.json"
    if not os.path.exists(json_path):
        # try direct API data
        medals_data = safe_json_load(json_path)
        if not medals_data:
            return []
        return medals_data

    medals = safe_json_load(json_path)
    rows = []
    for idx, m in enumerate(medals, 1):
        ruid = m.get("ruid", "")
        rows.append({
            "序号": idx,
            "粉丝牌名称": m.get("name", ""),
            "粉丝牌等级": m.get("level", 0),
            "主播UID": ruid,
            "主播空间": f"https://space.bilibili.com/{ruid}" if ruid else "",
        })
    return rows


# ═══════════════════════════════════════
# 3. 视频弹幕
# ═══════════════════════════════════════
def load_videodm():
    json_path = f"{OUT_DIR}/aicu_videodm_raw.json"
    dmlist = safe_json_load(json_path)
    if not dmlist:
        return []

    # 加载视频标题缓存
    cache_path = f"{CACHE_DIR}/aicu_video_titles_cache.json"
    titles = safe_json_load(cache_path, {})

    def ts_to_str(ts):
        if not ts: return ""
        from datetime import datetime, timezone, timedelta
        return datetime.fromtimestamp(ts, tz=timezone(timedelta(hours=8))).strftime("%Y-%m-%d %H:%M:%S")

    def ms_to_time(ms):
        if not ms: return ""
        s = int(ms) // 1000
        return f"{s//3600}:{(s%3600)//60:02d}:{s%60:02d}"

    rows = []
    for idx, dm in enumerate(dmlist, 1):
        oid = dm.get("oid", "")
        rows.append({
            "序号": idx,
            "弹幕ID": dm.get("id", ""),
            "弹幕内容": dm.get("content", ""),
            "发送时间": ts_to_str(dm.get("ctime", 0)),
            "视频ID": oid,
            "视频链接": f"https://www.bilibili.com/video/av{oid}" if oid else "",
            "视频标题": titles.get(str(oid), "") if oid else "",
            "出现时间点": ms_to_time(dm.get("progress", 0)),
            "进度(毫秒)": dm.get("progress", 0),
        })
    return rows


# ═══════════════════════════════════════
# 4. 直播弹幕
# ═══════════════════════════════════════
def load_livedm():
    json_path = f"{OUT_DIR}/aicu_livedm_raw.json"
    groups = safe_json_load(json_path)
    if not groups:
        return []

    def ts_to_str(ts):
        if not ts: return ""
        from datetime import datetime, timezone, timedelta
        return datetime.fromtimestamp(ts, tz=timezone(timedelta(hours=8))).strftime("%Y-%m-%d %H:%M:%S")

    rows = []
    idx = 0
    for group in groups:
        ri = group.get("roominfo", {})
        for dm in group.get("danmu", []):
            idx += 1
            rows.append({
                "序号": idx,
                "直播间ID": ri.get("roomid", ""),
                "主播昵称": ri.get("upname", ""),
                "主播UID": ri.get("upuid", ""),
                "直播间标题": ri.get("roomname", ""),
                "弹幕内容": dm.get("text", ""),
                "发送时间": ts_to_str(dm.get("ts", 0)),
            })
    return rows


# ═══════════════════════════════════════
# 5. 关键信息推断表（复用之前的逻辑）
# ═══════════════════════════════════════
def build_profile_sheet(ws, comment_rows):
    findings = [
        ("性别", "女性", "★★★★★",
         "自称「我一女的都觉得你糖欸」", [253], "第253行"),
        ("年龄", "约18-22岁", "★★★★",
         '问"是高考吗"；密集讨论中考总分；自称168cm/85斤', [2, 70, 71, 74, 75, 310], "第2/70/71/74/75/310行"),
        ("身高体重", "168cm / 42.5kg (85斤)", "★★★★★",
         "「丸辣，我168才85斤」直接自述", [310], "第310行"),
        ("籍贯", "云南省昆明市安宁市", "★★★★★",
         '"来昆明安宁看看"；洋芋叫芋头的滇中方言；描述昆明天气', [183, 184, 182], "第183/184/182行"),
        ("现居地", "可能在塞尔维亚（欧洲）", "★★★★",
         '"在境外呢"；回复"塞尔维亚"；23%评论在凌晨0-5点', [178, 237], "第178/237行+时段"),
        ("学历", "在读大学生", "★★★",
         '讨论高考/中考；"四个半小时一次性背完刑法"', [2, 70, 280], "第2/70/280行"),
        ("专业方向", "可能法律/计算机相关", "★★★",
         "背刑法；讨论水稻种质资源案；关注Python编程", [280, 124, 90], "第280/124/90行"),
        ("核心兴趣-王者荣耀", "上官婉儿专精玩家", "★★★★★",
         '自称"只会上官"；深度讨论T0-T5梯度、国服分、营地数据', [89, 144, 152, 155, 168, 213, 235, 247], "（代表样本）"),
        ("核心兴趣-Minecraft", "Java版生电技术向玩家", "★★★★★",
         '讨论TNT复制、刷沙机、满置域、史山代码、凋灵机制', [5, 6, 7, 9, 25, 42, 43, 52], "（代表样本）"),
        ("核心兴趣-音游", "Phigros/Arcaea 玩家", "★★★★",
         'Phigros相关评论；打出APT音游meme；"打音游的就没有几个成绩差的"', [251, 289, 293, 299], "第251/289/293/299行"),
        ("兴趣-围棋", "了解围棋规则", "★★★★",
         "讨论三劫循环、双活、柯洁争议", [198, 205, 206, 208], "第198/205/206/208行"),
        ("兴趣-小说", "看番茄小说/网文", "★★★",
         "了解《十日终焉》、番茄小说排行", [292], "第292行"),
        ("性格-理性数据派", "用数据说话", "★★★★★",
         '频繁引用"营地数据""国服分""胜率"论据', [152, 229, 247, 256, 257], "第152/229/247/256/257行"),
        ("性格-好辩但能认错", "会辩论也能承认错误", "★★★★",
         '长文争论后说"之前我确实不太理智了"', [225], "第225行"),
        ("性格-爱国但有判断", "爱国不盲从", "★★★★",
         '"看不到科技进步觉得不行去国外"；也指出黑神话缺点', [11, 301], "第11/301行"),
        ("性格-略带毒舌", "嘴不饶人", "★★★",
         '"钻石局的fvv又来评头论足了""主页没东西说话就是硬气"', [249, 191], "第249/191行"),
        ("价值观-成熟度", "社会判断力超年龄", "★★★★",
         '19岁生育："19岁就有孩子不负责""可怜但也可恨"', [77, 80, 81], "第77/80/81行"),
    ]

    # 标题
    ws.merge_cells('A1:G1')
    t1 = ws.cell(row=1, column=1, value=f'B站用户 UID {UID} 关键信息推断表')
    t1.font = Font(name='微软雅黑', bold=True, size=16, color='1A1A2E')
    t1.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    ws.merge_cells('A2:G2')
    t2 = ws.cell(row=2, column=1, value=f'数据来源：aicu.cc | 基于 {len(comment_rows)} 条评论 | 含推测成分仅供参考')
    t2.font = Font(name='微软雅黑', size=10, color='666666')
    t2.alignment = Alignment(horizontal='center', vertical='center')

    headers = ['类别', '推断结论', '置信度', '关键依据', '对应行号', '行号说明', '原始评论全文']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = HEADER_ALIGN; c.border = THIN_BORDER

    row_num = 5
    for item in findings:
        cat, conc, conf, ev, rns, rdesc = item

        ws.cell(row=row_num, column=1, value=cat).font = Font(name='微软雅黑', bold=True, size=11, color='1A1A2E')
        ws.cell(row=row_num, column=2, value=conc).font = Font(name='微软雅黑', bold=True, size=12, color='2B579A')
        ws.cell(row=row_num, column=3, value=conf).font = Font(name='微软雅黑', size=11)
        ws.cell(row=row_num, column=4, value=ev).font = Font(name='微软雅黑', size=10)
        ws.cell(row=row_num, column=5, value=', '.join(str(n) for n in rns)).font = Font(name='Consolas', size=11, color='C0392B', bold=True)
        ws.cell(row=row_num, column=6, value=rdesc).font = Font(name='微软雅黑', size=9, color='888888', italic=True)

        # 评论全文
        full = []
        for rn in rns:
            if 1 <= rn <= len(comment_rows):
                r = comment_rows[rn - 1]
                msg = r.get('评论内容', '').replace('\n', ' ')
                ts = r.get('评论时间', '')
                title = r.get('视频标题', '')
                full.append(f"[第{rn}行 | {ts}] {msg}")
                if title:
                    full.append(f"   -> {title}")
        ws.cell(row=row_num, column=7, value='\n'.join(full)).font = Font(name='微软雅黑', size=9, color='333333')

        for col in range(1, 8):
            c = ws.cell(row=row_num, column=col)
            c.alignment = CELL_ALIGN; c.border = THIN_BORDER
            if conf.count('★') >= 5:
                c.fill = HIGHLIGHT_FILL

        row_num += 1

    ws.column_dimensions['A'].width = 22; ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10; ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 20; ws.column_dimensions['F'].width = 28
    ws.column_dimensions['G'].width = 60
    ws.freeze_panes = 'A5'

    note_row = row_num + 2
    ws.merge_cells(f'A{note_row}:G{note_row}')
    ws.cell(row=note_row, column=1,
        value='说明：行号 = "评论数据"工作表中的序号列。黄色高亮 = 置信度五星（直接自述）。').font = Font(name='微软雅黑', size=9, color='999999', italic=True)


# ═══════════════════════════════════════
# 主程序
# ═══════════════════════════════════════
def main():
    wb = Workbook()

    # Sheet 1: 评论数据
    print("Loading comments...")
    comments = load_comments()
    ws1 = wb.active
    ws1.title = "评论数据"
    write_sheet(ws1, comments, {
        '序号': 6, '评论ID': 18, '评论内容': 50, '评论时间': 14, '层级': 10,
        '根评论ID': 16, '父评论ID': 16, '视频ID': 18, '视频链接': 38, '视频标题': 40, '评论跳转链接': 42,
    })
    print(f"  Comments: {len(comments)} rows")

    # Sheet 2: 视频弹幕
    print("Loading video danmaku...")
    videodm = load_videodm()
    ws2 = wb.create_sheet("视频弹幕")
    write_sheet(ws2, videodm, {
        '序号': 6, '弹幕ID': 22, '弹幕内容': 45, '发送时间': 18, '视频ID': 18,
        '视频链接': 38, '视频标题': 38, '出现时间点': 14, '进度(毫秒)': 14,
    })
    print(f"  Video Danmaku: {len(videodm)} rows")

    # Sheet 3: 直播弹幕
    print("Loading live danmaku...")
    livedm = load_livedm()
    ws3 = wb.create_sheet("直播弹幕")
    write_sheet(ws3, livedm, {
        '序号': 6, '直播间ID': 14, '主播昵称': 16, '主播UID': 16,
        '直播间标题': 35, '弹幕内容': 50, '发送时间': 18,
    })
    print(f"  Live Danmaku: {len(livedm)} rows")

    # Sheet 4: 粉丝牌
    print("Loading medals...")
    medals = load_medals()
    ws4 = wb.create_sheet("粉丝牌")
    write_sheet(ws4, medals, {
        '序号': 6, '粉丝牌名称': 16, '粉丝牌等级': 12, '主播UID': 16, '主播空间': 38,
    })
    print(f"  Medals: {len(medals)} rows")

    # Sheet 5: 关键信息推断
    if comments:
        print("Building profile sheet...")
        ws5 = wb.create_sheet("关键信息推断")
        build_profile_sheet(ws5, comments)
        print(f"  Profile: {18} inferences")

    # 保存
    out_path = f"{OUT_DIR}/aicu_data_{UID}.xlsx"
    wb.save(out_path)
    print(f"\n[DONE] Excel saved: {out_path}")
    print(f"  Sheets: {wb.sheetnames}")

if __name__ == "__main__":
    main()
