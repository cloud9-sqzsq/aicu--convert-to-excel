# -*- coding: utf-8 -*-
"""为 UID 93403197 生成含证据表格的分析 Excel"""

import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

UID = "93403197"
DATA_DIR = f"E:/aicu评论-ds-wb/output/{UID}"
OUT_PATH = f"{DATA_DIR}/aicu_analysis_{UID}.xlsx"

# ── 公共样式 ──
HDR_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
HDR_FILL = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_ALIGN = Alignment(vertical='top', wrap_text=True)
BORDER = Border(
    left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'),
)
GOLD = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
RED = PatternFill(start_color='FFF0F0', end_color='FFF0F0', fill_type='solid')
GREEN = PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid')

def write_sheet(ws, rows, widths=None):
    if not rows:
        ws.cell(row=1, column=1, value="(无数据)").font = Font(name='微软雅黑', size=12, color='999999')
        return
    fields = list(rows[0].keys())
    for ci, name in enumerate(fields, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = HDR_ALIGN; c.border = BORDER
    id_keywords = ('id', 'rpid', 'oid', 'rootid', 'parentid')
    def _is_id(name): return any(k in name.lower() for k in id_keywords)
    for ri, row in enumerate(rows, 2):
        for ci, name in enumerate(fields, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(name, ''))
            c.alignment = CELL_ALIGN; c.border = BORDER; c.font = Font(name='微软雅黑', size=10)
            if _is_id(name):
                c.number_format = '@'
    if widths:
        for ci, name in enumerate(fields, 1):
            ws.column_dimensions[get_column_letter(ci)].width = widths.get(name, 18)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{len(rows)+1}"


# ── 读取评论 ──
with open(f"{DATA_DIR}/aicu_comments_{UID}.csv", encoding='utf-8-sig') as f:
    comments = list(csv.DictReader(f))


# ═══════════════════════════════════════════
# 证据数据
# ═══════════════════════════════════════════
# (大类, 推断结论, 置信度, 依据摘要, 行号列表, 行号说明, 关键评论摘录)

evidence = [
    # === 基本信息 ===
    ("身份-北京人", "北京户籍，身份证 110 开头", "★★★★★",
     "多条评论直接或间接表明北京身份",
     [279, 90, 99, 286, 81, 180],
     "第279/90/99/286/81/180行",
     "「我身份证号110开头」(279);「110104孩子」(90);「我在北京生活的很好」(99);「叫声京爷我听听」(286);「北京公立学校比你们放得早」(81);「你就是被科学抛弃的人」(180)"),

    ("身份-北京人", "住家附近有大型家乐福超市（四环内）", "★★★",
     "「我家旁边有个大的」评论北京家乐福关店视频",
     [220],
     "第220行",
     "「我家旁边有个大的」(220)——关于家乐福关店视频"),

    ("身份-初中", "初中就读北京第十七中学", "★★★★",
     "直接提及初中学校",
     [45],
     "第45行",
     "「我初中也是十七中的 只不过不是这个校区」(45)"),

    ("年龄-高考生", "2025年北京高考生，物化生", "★★★★★",
     "密集讨论高考分数、科目，时间点为2025年6月",
     [10, 12, 13, 7, 8, 9, 19, 20],
     "第10/12/13/7/8/9/19/20行",
     "「我生物97 数学平常120考了个97」(12);「我也563物化生」(13);「英语一模69分俩月提到103」(10);「化学本来79水平这回才70」(11);「600排11000才能上211」(8);「bda ccda gedca听说的答案」(19)"),

    ("年龄-高考生", "2025年6月8-9日密集讨论刚结束的高考", "★★★★★",
     "高考英语第二天在相关视频下集中评论",
     [15, 16, 17, 18, 19, 20],
     "第15-20行",
     "6月8-9日密集发10+条高考相关评论，自称刚考完"),

    ("年龄-轨迹", "2019年6月就有评论（当时约12岁小学毕业）", "★★★★★",
     "评论时间跨度为2019-2025年，覆盖整个中学阶段",
     [569],
     "最后一行",
     "最早评论2019年6月13日，最早接触B站时约12岁"),

    ("健康-高血压", "高压169/低压90/心率175", "★★★★★",
     "直接自述体检数据",
     [102],
     "第102行",
     "「我体检测的高压169 低压90 心率175」(102)——在高血压科普视频下"),

    ("健康-高血压", "关注血压/心脏相关科普视频", "★★★★",
     "在多条医学健康视频下评论",
     [102, 226],
     "第102/226行",
     "「16.....」(226)——在「血压升高时心脏是怎么求救的」视频"),

    # === 性格特征 ===
    ("性格-攻击性", "84%的评论是回复别人（极高战斗倾向）", "★★★★★",
     "统计结论：569条中480条为'回复'",
     [],
     "全表统计",
     "569条评论，480条=回复(84.3%)，89条=一级评论(15.6%)"),

    ("性格-攻击性", "45次使用侮辱性词汇", "★★★★★",
     "统计数据",
     [42, 49, 56, 128, 166, 170, 173, 181, 235],
     "多处",
     "「跟狗做一桌吃饭去」(42);「滚」(56);「蠢是你一辈子的事」(128);「废物」(173);「nt」(235);「沙必」(122);「黄色废料姐」(49);「你就是被科学抛弃的人」(180)"),

    ("性格-优越感", "频繁以'京爷'身份碾压外地人", "★★★★★",
     "大量评论体现北京户籍优越感",
     [286, 76, 37, 93, 99, 149],
     "第286/76/37/93/99/149行",
     "「叫声京爷我听听」(286);「山东人指点上北京空气不好了」(76);「你IP还不配地域歧视」(37);「哪来的沙东人快去考公务员去」(93);「你在北京生活的很好 你呢」(99)"),

    ("性格-优越感", "频繁地域攻击：山东/河南/广东/江西/甘肃/四川", "★★★★★",
     "多个独立评论攻击不同省份",
     [93, 37, 149, 283, 129, 76, 116, 232, 176, 245],
     "第93/37/149/283/129/76/116/232/176/245行",
     "沙东人(93);河南中中中(283);老广吃死老鼠(129);山东人(76);那是你甘肃(116);江西人你好(232);满洲人(350);井盖人(245)"),

    ("性格-键政", "引用宪法条文对线", "★★★★★",
     "背诵《宪法》第四十一条全文来论证监督权",
     [104],
     "第104行",
     "「根据《中华人民共和国宪法》第四十一条...公民对于任何国家机关和国家工作人员，有提出批评和建议的权利」(104)——全文引用宪法"),

    ("性格-键政", "同时引用马列和哈耶克攻击不同立场的人", "★★★★★",
     "左右互搏式引经据典",
     [71, 68, 69, 167, 66, 67],
     "第71/68/69/167/66/67行",
     "「拿个几百年前的理论当真理」(71)——嘲讽剩余价值理论;「给你推荐《国家与gm》《反杜林论》」(68)——转头引用马列;「你读过经济学几本书就哈耶克上了」(167)"),

    ("性格-对话模式", "经典互联网吵架三板斧：'急了''典''看xx看的'", "★★★★★",
     "高频使用网络吵架定型语",
     [272, 273, 270, 269, 239, 247, 158, 173, 181],
     "第272/273/270/269/239/247/158行",
     "「急急急」(272);「哈哈急了」(273);「经典哑口无言」(269);「玩原神玩的」(239);「看二次元看的」(247);「玩火影玩的」(158);「你看你玩原神就知道你什么水平了」(181)"),

    ("性格-双标", "自己骂人，但反感别人'指点江山'", "★★★★",
     "高频批评别人'指点江山'，但自己正是在做什么",
     [186, 284],
     "第186/284行",
     "「指点江山的真的很烦」(186);「最烦的就是你们这种人 啥都不知道就指点江山」(284)"),

    # === 兴趣爱好 ===
    ("兴趣-手机数码", "深度关注手机/芯片/数码评测", "★★★★★",
     "68次关键词命中，能引用苹果官网原文",
     [210, 154, 147, 260, 174, 95, 97, 213],
     "第210/154/147/260/174/95/97/213行",
     "引用苹果官网M2 Ultra介绍全文(210);「8g2 GPU有750~1050水平」(260);「1050是垃圾」(147);「红米别碰瓷」(97);「人家512的是Pro」(213)"),

    ("兴趣-手机数码", "可能是苹果用户，维护苹果生态", "★★★★",
     "多条评论为苹果辩护/嘲讽安卓",
     [95, 255, 97, 213],
     "第95/255/97/213行",
     "「人家从一开始就不考虑安卓你比什么三家」(95);「果狗一个」(255)—别人骂果粉;「红米别碰瓷」(97)"),

    ("兴趣-二次元/游戏", "反原神/反二次元鄙视链执行者", "★★★★★",
     "63次相关关键词，以嘲讽为主",
     [239, 247, 173, 158, 160, 181, 230],
     "第239/247/173/158/160/181/230行",
     "「玩原神玩的」(239);「看二次元看的」(247);「玩原神的有点儿废物了」(173);「皱皮又懂了」(161);「二次元青蛙滚」(178);「你看你玩原神就知道你什么水平了」(181)"),

    ("兴趣-政治经济讨论", "关注经济发展/德国历史/乒乓/AI/转基因", "★★★★",
     "涉猎广泛但不深入",
     [128, 148, 119, 291, 158, 150],
     "第128/148/119/291/150行",
     "讨论两德统一(128);讨论GDP(148);讨论转基因(119);讨论ChatGPT(291);AI绘画著作权(150)"),

    # === 时间模式 ===
    ("行为-时间", "凌晨活跃比例低（6%），作息相对正常", "★★★",
     "凌晨0-5点评论仅6%",
     [],
     "时段统计",
     "39/569 条评论在凌晨0-5点(6.9%)，远低于上一位的23%"),
]

# ═══════════════════════════════════════════
# 构建 Excel
# ═══════════════════════════════════════════
wb = Workbook()

# --- Sheet 1: 评论数据 ---
ws1 = wb.active
ws1.title = "评论数据"
write_sheet(ws1, comments, {
    '序号': 6, '评论ID': 18, '评论内容': 55, '评论时间': 14, '层级': 10,
    '根评论ID': 16, '父评论ID': 16, '视频ID': 18, '视频链接': 38, '视频标题': 40,
    '评论跳转链接': 42,
})

# 高亮关键证据行
key_rows = {279, 90, 99, 286, 81, 180, 45, 10, 12, 13, 7, 8, 9, 19, 20,
            102, 42, 49, 56, 128, 166, 170, 173, 181, 235, 76, 37, 93, 149,
            283, 129, 116, 232, 176, 245, 104, 71, 68, 69, 167, 66, 67,
            272, 273, 270, 269, 239, 247, 158, 186, 284, 210, 154, 147, 260,
            95, 97, 213, 255, 161, 178, 150, 148, 119, 291}
for rn in key_rows:
    if 1 <= rn <= len(comments):
        for ci in range(1, 13):
            ws1.cell(row=rn+1, column=ci).fill = PatternFill(start_color='FFFFF0', end_color='FFFFF0', fill_type='solid')

# --- Sheet 2: 关键证据表 ---
ws2 = wb.create_sheet("关键信息推断（含证据）")

# 标题
ws2.merge_cells('A1:H1')
t = ws2.cell(row=1, column=1, value=f'B站用户 UID {UID} 关键信息推断表（含证据行号）')
t.font = Font(name='微软雅黑', bold=True, size=16, color='1A1A2E')
t.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 36

ws2.merge_cells('A2:H2')
t = ws2.cell(row=2, column=1, value=f'数据来源：aicu.cc | 569条评论 | 2019~2025 | 每条推断均可按行号在"评论数据"表中验证')
t.font = Font(name='微软雅黑', size=10, color='666666')
t.alignment = Alignment(horizontal='center')

cols = ['类别', '推断结论', '置信度', '依据摘要', '对应行号', '行号说明', '关键评论摘录', '原始评论全文']
for ci, h in enumerate(cols, 1):
    c = ws2.cell(row=4, column=ci, value=h)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = HDR_ALIGN; c.border = BORDER

rn = 5
prev_cat = ''
for item in evidence:
    cat, conc, conf, summary, row_nums, row_desc, excerpts = item

    # 类别
    c = ws2.cell(row=rn, column=1, value=cat)
    c.font = Font(name='微软雅黑', bold=True, size=11, color='1A1A2E')
    c.alignment = CELL_ALIGN; c.border = BORDER

    # 结论
    c = ws2.cell(row=rn, column=2, value=conc)
    c.font = Font(name='微软雅黑', bold=True, size=12, color='2B579A')
    c.alignment = CELL_ALIGN; c.border = BORDER

    # 置信度
    c = ws2.cell(row=rn, column=3, value=conf)
    c.font = Font(name='微软雅黑', size=11)
    c.alignment = Alignment(horizontal='center', vertical='top'); c.border = BORDER

    # 依据摘要
    c = ws2.cell(row=rn, column=4, value=summary)
    c.font = Font(name='微软雅黑', size=10); c.alignment = CELL_ALIGN; c.border = BORDER

    # 行号
    nums = ', '.join(str(x) for x in row_nums) if row_nums else '—'
    c = ws2.cell(row=rn, column=5, value=nums)
    c.font = Font(name='Consolas', size=11, color='C0392B', bold=True)
    c.alignment = Alignment(horizontal='center', vertical='top'); c.border = BORDER

    # 行号说明
    c = ws2.cell(row=rn, column=6, value=row_desc)
    c.font = Font(name='微软雅黑', size=9, color='888888', italic=True)
    c.alignment = CELL_ALIGN; c.border = BORDER

    # 关键摘录
    c = ws2.cell(row=rn, column=7, value=excerpts)
    c.font = Font(name='微软雅黑', size=10, color='333333'); c.alignment = CELL_ALIGN; c.border = BORDER

    # 原始评论全文
    full = []
    for rn_idx in row_nums:
        if 1 <= rn_idx <= len(comments):
            r = comments[rn_idx - 1]
            msg = r['评论内容'].replace('\n', ' ')
            ts = r['评论时间']
            title = r.get('视频标题', '')
            full.append(f"[第{rn_idx}行 | {ts}] {msg}")
            if title and '获取失败' not in title and '不存在' not in title:
                full.append(f"  -> 视频: {title}")
    c = ws2.cell(row=rn, column=8, value='\n'.join(full) if full else summary)
    c.font = Font(name='微软雅黑', size=9, color='555555'); c.alignment = CELL_ALIGN; c.border = BORDER

    # 高亮
    stars = conf.count('★')
    if stars >= 5:
        for col in range(1, 9):
            ws2.cell(row=rn, column=col).fill = GOLD
    elif stars == 4:
        pass  # 白色
    else:
        pass

    rn += 1

# 列宽
ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 32
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 40
ws2.column_dimensions['E'].width = 15
ws2.column_dimensions['F'].width = 22
ws2.column_dimensions['G'].width = 55
ws2.column_dimensions['H'].width = 65
ws2.freeze_panes = 'A5'

# 说明行
nr = rn + 2
ws2.merge_cells(f'A{nr}:H{nr}')
ws2.cell(row=nr, column=1,
    value='说明：黄色高亮=⭐⭐⭐⭐⭐置信度（直接自述或统计数据）。行号=「评论数据」工作表中的序号列。请切换到Sheet1按行号验证原始评论。').font = Font(name='微软雅黑', size=9, color='999999', italic=True)

# --- Sheet 3: 统计摘要 ---
ws3 = wb.create_sheet("统计摘要")
ws3.merge_cells('A1:C1')
ws3.cell(row=1, column=1, value=f'UID {UID} 数据统计').font = Font(name='微软雅黑', bold=True, size=14)

stats = [
    ('评论总数', '569', ''),
    ('一级评论', '89 (15.6%)', '主动发言较少'),
    ('回复评论', '480 (84.3%)', '极度好斗，到处找人吵架'),
    ('时间跨度', '2019-06-13 ~ 2025-08-25', '覆盖6年，整个中学阶段'),
    ('涉及视频数', '394', '话题极为分散'),
    ('凌晨活跃(0-5点)', '39条 (6.9%)', '作息相对正常'),
    ('', '', ''),
    ('# 身份标签', '', ''),
    ('户籍', '北京 (身份证110开头)', '行279'),
    ('年龄', '18-19岁 (2025年高考)', '行10/12/13'),
    ('选科', '物化生', '行13'),
    ('高考总分', '563分', '行13'),
    ('初中', '北京第十七中学', '行45'),
    ('健康', '高压169/低压90/心率175', '行102'),
    ('', '', ''),
    ('# 行为标签', '', ''),
    ('侮辱性词汇', '45次', '统计'),
    ('地域攻击对象', '山东/河南/广东/江西/甘肃/四川', '统计'),
    ('键政引用', '宪法/马列/哈耶克均有引用', '行104/68/167'),
    ('吵架标志语', '"急了""典""玩原神玩的"', '统计'),
    ('数码兴趣', '苹果/华为/高通/芯片制程', '行210/260'),
]

for i, (k, v, note) in enumerate(stats, 3):
    ws3.cell(row=i, column=1, value=k).font = Font(name='微软雅黑', bold=True, size=11)
    ws3.cell(row=i, column=2, value=v).font = Font(name='微软雅黑', size=11)
    ws3.cell(row=i, column=3, value=note).font = Font(name='微软雅黑', size=10, color='888888')

ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 35
ws3.column_dimensions['C'].width = 40

# 保存
wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
print(f"Sheet1: 评论数据 ({len(comments)}行)")
print(f"Sheet2: 关键信息推断 ({len(evidence)}条)")
print(f"Sheet3: 统计摘要")
