# -*- coding: utf-8 -*-
"""为 UID 3546630485707666 生成分析 Excel（仅3条评论）"""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

UID = "3546630485707666"
DATA_DIR = f"E:/aicu评论-ds-wb/output/{UID}"
OUT_PATH = f"{DATA_DIR}/aicu_analysis_{UID}.xlsx"

HDR_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
HDR_FILL = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_ALIGN = Alignment(vertical='top', wrap_text=True)
BORDER = Border(left=Side(style='thin',color='DDDDDD'),right=Side(style='thin',color='DDDDDD'),
                top=Side(style='thin',color='DDDDDD'),bottom=Side(style='thin',color='DDDDDD'))
GOLD = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')

def write_sheet(ws, rows, widths=None):
    if not rows: return
    fields = list(rows[0].keys())
    for ci, name in enumerate(fields, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = HDR_ALIGN; c.border = BORDER
    id_keywords = ('id', 'rpid', 'oid', 'rootid', 'parentid')
    def _is_id(name): return any(k in name.lower() for k in id_keywords)
    for ri, row in enumerate(rows, 2):
        for ci, name in enumerate(fields, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(name,''))
            c.alignment = CELL_ALIGN; c.border = BORDER; c.font = Font(name='微软雅黑', size=10)
            if _is_id(name):
                c.number_format = '@'
    if widths:
        for ci, name in enumerate(fields, 1):
            ws.column_dimensions[get_column_letter(ci)].width = widths.get(name, 18)
    ws.freeze_panes = 'A2'

with open(f"{DATA_DIR}/aicu_comments_{UID}.csv", encoding='utf-8-sig') as f:
    comments = list(csv.DictReader(f))

evidence = [
    ("行为-复制粘贴", "同一条200+字评论一字不改粘贴到两个不同视频下", "★★★★★",
     "行1和行2内容完全一致，仅视频不同（一个是「高级词汇帮作文50+」、一个是「中考满分作文高级词汇」）",
     [1, 2],
     "第1/2行",
     "两条评论内容100%相同：'感恩老师，我一定砥砺追忆您的昭示...'"),

    ("兴趣-应试写作", "关注中考/高考作文词汇视频", "★★★★★",
     "两条评论都在作文词汇教学视频下，一条涉及高考作文、一条涉及中考",
     [1, 2],
     "第1/2行",
     "[行1] 视频：「好歹毒的高级词！用了以后作文再也没下过50+」; [行2] 视频：「中考真实满分作文的高级词汇」"),

    ("年龄-推测", "约14-16岁（初三到高一阶段，在准备中高考）", "★★★",
     "同时关注中考和高考作文词汇，说明处于升学过渡期；用词生硬堆砌，是典型的应试词汇练习痕迹",
     [1, 2],
     "第1/2行",
     "评论中密集使用生僻词（倥偬/吊诡/荡漾/谙悉/涤荡/踯躅/逼仄/阡陌/倏忽/韵华），明显在练习高级词汇"),

    ("行为-刷题式评论", "把评论区当作词汇练习本，而非真实社交互动", "★★★★",
     "评论内容是一段塞满20+个生僻词汇的堆砌段落，像是完成了「用这些词造句」的作业",
     [1],
     "第1行",
     "200+字的评论本质上是词汇填空题，没有任何针对视频内容的真实回应"),

    ("时间-活跃", "仅3条评论，集中在2025年1月底至2月初（寒假期间）", "★★★★",
     "2025年1月26日~2月2日，三条评论跨一周。寒假期间有空刷B站并做词汇练习",
     [1, 2, 3],
     "第1/2/3行",
     "三条评论时间：2025-01-26, 2025-02-02(两条同天)"),

    ("性格-伪成熟", "用远超年龄的词汇量包装浅薄的内容，典型的考场作文习气", "★★★★",
     "堆砌了倥偬、吊诡、荡漾、谙悉等20+生僻词，但内容空洞，看完不知道在说什么",
     [1, 2],
     "第1/2行",
     "全文核心信息=0，信息密度=0，词汇密度=过量。标准的「凑字数凑高级词」考场套路"),
]

# ═══════ Excel ═══════
wb = Workbook()

ws1 = wb.active
ws1.title = "评论数据"
write_sheet(ws1, comments, {
    '序号': 6, '评论ID': 18, '评论内容': 65, '评论时间': 14, '层级': 10,
    '根评论ID': 16, '父评论ID': 16, '视频ID': 18, '视频链接': 38, '视频标题': 40,
    '评论跳转链接': 45,
})

ws2 = wb.create_sheet("关键信息推断（含证据）")
ws2.merge_cells('A1:H1')
ws2.cell(row=1, column=1, value=f'B站用户 UID {UID} 个人信息推断表（样本量极少，仅3条）').font = Font(name='微软雅黑', bold=True, size=16, color='1A1A2E')
ws2.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 36

ws2.merge_cells('A2:H2')
ws2.cell(row=2, column=1, value=f'⚠ 仅3条评论(2025-01~02)，推断置信度有限 | 数据来源：aicu.cc').font = Font(name='微软雅黑', size=10, color='666666')
ws2.cell(row=2, column=1).alignment = Alignment(horizontal='center')

cols = ['类别', '推断结论', '置信度', '依据摘要', '对应行号', '行号说明', '关键评论摘录', '原始评论全文']
for ci, h in enumerate(cols, 1):
    c = ws2.cell(row=4, column=ci, value=h)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = HDR_ALIGN; c.border = BORDER

rn = 5
for item in evidence:
    cat, conc, conf, summary, row_nums, row_desc, excerpts = item

    ws2.cell(row=rn, column=1, value=cat).font = Font(name='微软雅黑', bold=True, size=11, color='1A1A2E')
    ws2.cell(row=rn, column=2, value=conc).font = Font(name='微软雅黑', bold=True, size=12, color='2B579A')
    c3 = ws2.cell(row=rn, column=3, value=conf)
    c3.font = Font(name='微软雅黑', size=11); c3.alignment = Alignment(horizontal='center', vertical='top')
    ws2.cell(row=rn, column=4, value=summary).font = Font(name='微软雅黑', size=10)
    nums = ', '.join(str(x) for x in row_nums) if row_nums else '—'
    c5 = ws2.cell(row=rn, column=5, value=nums)
    c5.font = Font(name='Consolas', size=11, color='C0392B', bold=True); c5.alignment = Alignment(horizontal='center', vertical='top')
    ws2.cell(row=rn, column=6, value=row_desc).font = Font(name='微软雅黑', size=9, color='888888', italic=True)
    ws2.cell(row=rn, column=7, value=excerpts).font = Font(name='微软雅黑', size=10, color='333333')

    full = []
    for rn_idx in row_nums:
        if 1 <= rn_idx <= len(comments):
            r = comments[rn_idx-1]
            msg = r['评论内容'].replace('\n',' ')
            ts = r['评论时间']; title = r.get('视频标题','')
            full.append(f"[第{rn_idx}行 | {ts}] {msg}")
            if title and '失败' not in title and '不存在' not in title:
                full.append(f"  -> {title}")
    ws2.cell(row=rn, column=8, value='\n'.join(full) if full else summary).font = Font(name='微软雅黑', size=9, color='555555')

    for col in range(1, 9):
        c = ws2.cell(row=rn, column=col)
        c.alignment = CELL_ALIGN; c.border = BORDER
        if conf.count('★') >= 5: c.fill = GOLD

    rn += 1

ws2.column_dimensions['A'].width = 20; ws2.column_dimensions['B'].width = 35
ws2.column_dimensions['C'].width = 10; ws2.column_dimensions['D'].width = 45
ws2.column_dimensions['E'].width = 12; ws2.column_dimensions['F'].width = 22
ws2.column_dimensions['G'].width = 55; ws2.column_dimensions['H'].width = 65
ws2.freeze_panes = 'A5'

wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
print(f"Comments: {len(comments)} | Inferences: {len(evidence)}")
