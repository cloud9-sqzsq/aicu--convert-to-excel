#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AICU.cc B站数据爬虫 - 图形化界面
===============================
基于 aicu.cc API，一键爬取用户的评论、视频弹幕、直播弹幕、粉丝牌，
输出为 Excel 多表文件。支持自定义输出目录和 UID 输入。

API 端点:
  /api/v3/search/getreply   - 评论
  /api/v3/search/getvideodm  - 视频弹幕
  /api/v3/search/getlivedm   - 直播弹幕
  /api/v3/user/getmedal      - 粉丝牌

* B站评论 IP 属地和用户动态需要登录态，aicu.cc 暂不支持，工具中已标注。

依赖: pip install requests openpyxl curl_cffi
"""

import csv, json, os, sys, time, threading, traceback
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from datetime import datetime, timezone, timedelta
from collections import Counter

import requests
try:
    from curl_cffi import requests as cf_requests
    HAS_CFFI = True
except ImportError:
    HAS_CFFI = False

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 配置 ───────────────────────────────────────────
AICU_BASE = "https://api.aicu.cc/api/v3"
BILI_VIDEO_INFO = "https://api.bilibili.com/x/web-interface/view"
BILI_VIDEO_URL = "https://www.bilibili.com/video/av{oid}"
DEFAULT_PS = 500
REQUEST_DELAY = 0.6
BILI_DELAY = 0.3
TIMEOUT = 30
MAX_RETRIES = 3
RETRY_BACKOFF = 2
BEIJING_TZ = timezone(timedelta(hours=8))
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36")

# ── Excel 样式 ─────────────────────────────────────
HDR_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
HDR_FILL = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_ALIGN = Alignment(vertical='top', wrap_text=True)
BORDER = Border(left=Side(style='thin',color='DDDDDD'), right=Side(style='thin',color='DDDDDD'),
                top=Side(style='thin',color='DDDDDD'), bottom=Side(style='thin',color='DDDDDD'))


# ═══════════════════════════════════════════════════
# 工具函数
# ═══════════════════════════════════════════════════
def ts_beijing(ts):
    if not ts: return ""
    if ts > 1e12: ts //= 1000
    return datetime.fromtimestamp(ts, tz=BEIJING_TZ).strftime("%Y-%m-%d %H:%M:%S")

def oid_to_url(oid):
    return BILI_VIDEO_URL.format(oid=oid) if oid else ""

def comment_jump(oid, rpid):
    return f"https://www.bilibili.com/video/av{oid}#reply{rpid}" if oid and rpid else ""

def ms_to_time(ms):
    if not ms: return ""
    s = int(ms) // 1000
    return f"{s//3600}:{(s%3600)//60:02d}:{s%60:02d}"

def safe_print(text):
    """防 GBK 编码崩溃的 print"""
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode('ascii', errors='replace').decode('ascii'))

# ── ID 格式保护 ────────────────────────────────────
ID_KEYWORDS = ('id', 'rpid', 'oid', 'rootid', 'parentid', 'uid', 'ruid', 'roomid')
def _is_id_col(name):
    return any(k in name.lower() for k in ID_KEYWORDS)


# ═══════════════════════════════════════════════════
# 通用 HTTP 请求
# ═══════════════════════════════════════════════════
def api_get(endpoint, params):
    url = f"{AICU_BASE}{endpoint}"
    headers = {"User-Agent": UA, "Referer": "https://www.aicu.cc/"}
    for attempt in range(MAX_RETRIES):
        try:
            if HAS_CFFI:
                resp = cf_requests.get(url, params=params, headers=headers,
                                       timeout=TIMEOUT, impersonate="chrome124")
            else:
                resp = requests.get(url, params=params, headers=headers, timeout=TIMEOUT)
            if resp.status_code in (403, 468):
                if attempt < MAX_RETRIES - 1:
                    time.sleep(RETRY_BACKOFF ** (attempt + 1)); continue
            resp.raise_for_status()
            data = resp.json()
            if data.get("code") != 0:
                if attempt < MAX_RETRIES - 1:
                    time.sleep(RETRY_BACKOFF ** attempt); continue
            return data
        except Exception as e:
            if attempt < MAX_RETRIES - 1: time.sleep(RETRY_BACKOFF ** attempt)
            else: return {"code": -1, "error": str(e)}
    return {"code": -1}


# ═══════════════════════════════════════════════════
# 视频标题
# ═══════════════════════════════════════════════════
def fetch_titles(oids, cache_dir, log_func=print):
    sep = os.path.join(cache_dir, "aicu_video_titles_cache.json")
    cache = {}
    if os.path.exists(sep):
        with open(sep, encoding='utf-8') as f: cache = json.load(f)
    uncached = [o for o in oids if str(o) not in cache and o]
    if not uncached:
        log_func(f"  视频标题: 全部命中缓存 ({len(oids)} 个)")
        return {str(k): v for k, v in cache.items()}
    log_func(f"  视频标题: 需查询 {len(uncached)}/{len(oids)}")
    for i, oid in enumerate(uncached, 1):
        title = "(获取失败)"
        for _ in range(MAX_RETRIES):
            try:
                resp = requests.get(BILI_VIDEO_INFO, params={"aid": oid},
                                    timeout=TIMEOUT, headers={"User-Agent": UA})
                if resp.status_code == 412: time.sleep(3); continue
                d = resp.json()
                title = d["data"]["title"] if d.get("code") == 0 else ("(视频不存在)" if d.get("code")==-404 else f"(API:{d.get('code')})")
                break
            except: pass
        cache[str(oid)] = title
        if i % 20 == 0:
            log_func(f"    标题: {i}/{len(uncached)}")
        if i % 10 == 0:
            with open(sep, 'w', encoding='utf-8') as f: json.dump(cache, f, ensure_ascii=False, indent=2)
        time.sleep(BILI_DELAY)
    with open(sep, 'w', encoding='utf-8') as f: json.dump(cache, f, ensure_ascii=False, indent=2)
    return {str(k): v for k, v in cache.items()}


# ═══════════════════════════════════════════════════
# 爬取函数
# ═══════════════════════════════════════════════════
def crawl_reply(uid, log_func=print):
    """评论"""
    log_func("[评论] 开始爬取...")
    all_items, page = [], 1
    first = api_get("/search/getreply", {"uid": uid, "pn": 1, "ps": DEFAULT_PS, "mode": 0})
    if first.get("code") != 0: return []
    data = first["data"]
    total = data["cursor"]["all_count"]
    all_items.extend(data["replies"])
    log_func(f"  总评论: {total}  页数: {(total-1)//DEFAULT_PS+1}")
    is_end = data["cursor"]["is_end"]
    time.sleep(REQUEST_DELAY)
    while not is_end and len(all_items) < total:
        page += 1
        resp = api_get("/search/getreply", {"uid": uid, "pn": page, "ps": DEFAULT_PS, "mode": 0})
        if resp.get("code") != 0: continue
        d = resp["data"]
        all_items.extend(d["replies"])
        is_end = d["cursor"]["is_end"]
        log_func(f"  第{page}页 {len(d['replies'])}条 | 累计{len(all_items)}/{total}")
        time.sleep(REQUEST_DELAY)
    log_func(f"[评论] 完成: {len(all_items)} 条")
    return all_items

def crawl_videodm(uid, log_func=print):
    """视频弹幕"""
    log_func("[视频弹幕] 开始爬取...")
    all_items, page = [], 1
    first = api_get("/search/getvideodm", {"uid": uid, "pn": 1, "ps": DEFAULT_PS})
    if first.get("code") != 0: return []
    data = first["data"]
    total = data["cursor"]["all_count"]
    all_items.extend(data["videodmlist"])
    log_func(f"  总弹幕: {total}")
    is_end = data["cursor"]["is_end"]
    time.sleep(REQUEST_DELAY)
    while not is_end and len(all_items) < total:
        page += 1
        resp = api_get("/search/getvideodm", {"uid": uid, "pn": page, "ps": DEFAULT_PS})
        if resp.get("code") != 0: continue
        d = resp["data"]
        all_items.extend(d["videodmlist"])
        is_end = d["cursor"]["is_end"]
        log_func(f"  第{page}页 | 累计{len(all_items)}/{total}")
        time.sleep(REQUEST_DELAY)
    log_func(f"[视频弹幕] 完成: {len(all_items)} 条")
    return all_items

def crawl_livedm(uid, log_func=print):
    """直播弹幕"""
    log_func("[直播弹幕] 开始爬取...")
    page, groups = 1, []
    first = api_get("/search/getlivedm", {"uid": uid, "pn": 1, "ps": 50})
    if first.get("code") != 0: return []
    d = first["data"]
    total = d["cursor"]["all_count"]
    groups.extend(d["list"])
    log_func(f"  总弹幕组: {total}")
    is_end = d["cursor"]["is_end"]
    time.sleep(REQUEST_DELAY)
    while not is_end:
        page += 1
        resp = api_get("/search/getlivedm", {"uid": uid, "pn": page, "ps": 50})
        if resp.get("code") != 0: continue
        d = resp["data"]
        groups.extend(d["list"])
        is_end = d["cursor"]["is_end"]
        log_func(f"  第{page}页 | 组{len(groups)}/{total}")
        time.sleep(REQUEST_DELAY)
    # 展平
    flat, idx = [], 0
    for g in groups:
        ri = g.get("roominfo", {})
        for dm in g.get("danmu", []):
            idx += 1
            flat.append(dict(序号=idx, 直播间ID=ri.get("roomid",""), 主播昵称=ri.get("upname",""),
                            主播UID=ri.get("upuid",""), 直播间标题=ri.get("roomname",""),
                            弹幕内容=dm.get("text",""), 发送时间=ts_beijing(dm.get("ts",0))))
    log_func(f"[直播弹幕] 完成: {len(flat)} 条")
    return flat

def crawl_medal(uid, log_func=print):
    """粉丝牌"""
    log_func("[粉丝牌] 开始爬取...")
    resp = api_get("/user/getmedal", {"uid": uid})
    if resp.get("code") != 0: return []
    medals = resp["data"]["list"]
    rows = []
    for i, m in enumerate(medals, 1):
        ruid = m.get("ruid","")
        rows.append(dict(序号=i, 粉丝牌名称=m.get("name",""), 粉丝牌等级=m.get("level",0),
                        主播UID=ruid, 主播空间=f"https://space.bilibili.com/{ruid}" if ruid else ""))
    log_func(f"[粉丝牌] 完成: {len(rows)} 个")
    return rows


# ═══════════════════════════════════════════════════
# 数据整理
# ═══════════════════════════════════════════════════
def flatten_reply(comments, titles, ip_map=None):
    rows = []
    for idx, c in enumerate(comments, 1):
        rpid = c.get("rpid",""); oid = c.get("dyn",{}).get("oid","")
        ts = c.get("time",0); parent = c.get("parent",{})
        ip = ""
        if ip_map and rpid in ip_map:
            ip = ip_map[rpid]
        elif not ip_map:
            ip = "(未登录)"
        rows.append(dict(
            序号=idx, 评论ID=rpid, 评论内容=c.get("message",""),
            评论时间=ts_beijing(ts), 层级="一级评论" if c.get("rank")==1 else "回复",
            根评论ID=parent.get("rootid",""), 父评论ID=parent.get("parentid",""),
            视频ID=oid, 视频链接=oid_to_url(oid),
            视频标题=titles.get(str(oid),"") if oid else "",
            评论跳转链接=comment_jump(oid, rpid),
            IP属地=ip or "(未登录)"
        ))
    return rows


def resolve_ips(reply_raw, sessdata, log_func=print):
    """通过B站API获取评论IP属地"""
    if not sessdata or not reply_raw: return {}
    log_func("\n[IP属地] 正在通过B站API获取...")
    ip_map = {}
    cookie = f"SESSDATA={sessdata}"
    cnt = 0
    for c in reply_raw:
        rpid = c.get("rpid","")
        oid = c.get("dyn",{}).get("oid","")
        if not rpid or not oid: continue
        try:
            resp = requests.get("https://api.bilibili.com/x/v2/reply/detail",
                                params={"oid": oid, "rpid": rpid, "type": 1},
                                headers={"User-Agent": UA, "Referer": "https://www.bilibili.com/",
                                         "Cookie": cookie}, timeout=10)
            d = resp.json()
            if d.get("code") == 0:
                reply = d["data"].get("reply")
                if reply:
                    ip_map[rpid] = reply.get("reply_control", {}).get("location", "未知")
            else:
                # 可能触发频率限制
                if d.get("code") == -799:
                    log_func("  [IP] 请求过于频繁，暂停5秒...")
                    time.sleep(5)
                continue
        except: pass
        cnt += 1
        if cnt % 20 == 0:
            log_func(f"  [IP] 进度: {cnt}/{len(reply_raw)}")
        time.sleep(0.2)
    log_func(f"[IP属地] 完成: {len(ip_map)} 条解析")
    return ip_map

def flatten_videodm(dms, titles):
    rows = []
    for idx, dm in enumerate(dms, 1):
        oid = dm.get("oid","")
        rows.append(dict(
            序号=idx, 弹幕ID=dm.get("id",""), 弹幕内容=dm.get("content",""),
            发送时间=ts_beijing(dm.get("ctime",0)), 视频ID=oid,
            视频链接=oid_to_url(oid),
            视频标题=titles.get(str(oid),"") if oid else "",
            出现时间点=ms_to_time(dm.get("progress",0)), 进度毫秒=dm.get("progress",0),
        ))
    return rows


# ═══════════════════════════════════════════════════
# Excel 生成
# ═══════════════════════════════════════════════════
def write_sheet(ws, rows, widths=None):
    if not rows: return
    fields = list(rows[0].keys())
    for ci, name in enumerate(fields, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = HDR_ALIGN; c.border = BORDER
    for ri, row in enumerate(rows, 2):
        for ci, name in enumerate(fields, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(name,''))
            c.alignment = CELL_ALIGN; c.border = BORDER; c.font = Font(name='微软雅黑', size=10)
            if _is_id_col(name): c.number_format = '@'
    if widths:
        for ci, name in enumerate(fields, 1):
            ws.column_dimensions[get_column_letter(ci)].width = widths.get(name, 18)
    ws.freeze_panes = 'A2'
    if len(rows) > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{len(rows)+1}"

def build_excel(uid, reply_rows, videodm_rows, livedm_rows, medal_rows, out_dir):
    wb = Workbook()
    reply_w = {'序号':6, '评论ID':18, '评论内容':55, '评论时间':14, '层级':10,
               '根评论ID':18, '父评论ID':18, '视频ID':18, '视频链接':38,
               '视频标题':38, '评论跳转链接':42, 'IP属地':12}
    videodm_w = {'序号':6, '弹幕ID':22, '弹幕内容':50, '发送时间':14, '视频ID':18,
                 '视频链接':38, '视频标题':38, '出现时间点':14, '进度毫秒':14}
    livedm_w = {'序号':6, '直播间ID':14, '主播昵称':16, '主播UID':16,
                '直播间标题':35, '弹幕内容':50, '发送时间':14}
    medal_w = {'序号':6, '粉丝牌名称':16, '粉丝牌等级':12, '主播UID':18, '主播空间':38}

    sheets = [
        ("评论数据", reply_rows, reply_w),
        ("视频弹幕", videodm_rows, videodm_w),
        ("直播弹幕", livedm_rows, livedm_w),
        ("粉丝牌", medal_rows, medal_w),
    ]
    first = True
    for name, rows, widths in sheets:
        if first:
            ws = wb.active; ws.title = name; first = False
        else:
            ws = wb.create_sheet(name)
        write_sheet(ws, rows, widths)

    path = os.path.join(out_dir, f"aicu_data_{uid}.xlsx")
    wb.save(path)
    return path


# ═══════════════════════════════════════════════════
# GUI
# ═══════════════════════════════════════════════════
class AicuApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("AICU 评论爬虫 - GUI版")
        self.root.geometry("720x750")
        self.root.resizable(True, True)
        self.running = False
        self.out_dir = tk.StringVar(value=os.path.abspath("output"))
        self.sessdata = tk.StringVar(value="")
        self._build_ui()

    def _build_ui(self):
        pad = dict(padx=10, pady=4)

        # ── 标题栏 ──
        title = tk.Label(self.root, text="AICU.cc B站用户数据爬虫",
                         font=("微软雅黑", 16, "bold"), fg="#1A1A2E")
        title.pack(**pad, pady=(12, 2))
        subtitle = tk.Label(self.root, text="评论 · 视频弹幕 · 直播弹幕 · 粉丝牌  |  数据来源: aicu.cc",
                            font=("微软雅黑", 9), fg="#888")
        subtitle.pack(**pad)

        # ── 输入区 ──
        inp_frame = tk.LabelFrame(self.root, text="查询设置", font=("微软雅黑", 11, "bold"),
                                  fg="#2B579A", padx=10, pady=8)
        inp_frame.pack(fill="x", **pad, pady=(10, 4))

        # UID
        row1 = tk.Frame(inp_frame); row1.pack(fill="x", pady=2)
        tk.Label(row1, text="B站 UID：", font=("微软雅黑", 10), width=10, anchor="e").pack(side="left")
        self.uid_entry = tk.Entry(row1, font=("Consolas", 13), width=22)
        self.uid_entry.pack(side="left", padx=8)
        self.uid_entry.focus()
        self.uid_entry.bind("<Return>", lambda e: self._start())

        # 输出目录
        row2 = tk.Frame(inp_frame); row2.pack(fill="x", pady=2)
        tk.Label(row2, text="输出目录：", font=("微软雅黑", 10), width=10, anchor="e").pack(side="left")
        self.dir_label = tk.Label(row2, textvariable=self.out_dir, font=("微软雅黑", 9),
                                  fg="#555", bg="#f5f5f5", relief="sunken",
                                  anchor="w", width=50)
        self.dir_label.pack(side="left", padx=8, fill="x", expand=True)
        tk.Button(row2, text="浏览...", command=self._browse_dir,
                  font=("微软雅黑", 9), width=8).pack(side="left", padx=4)

        # B站 登录
        row_login = tk.Frame(inp_frame); row_login.pack(fill="x", pady=4)
        tk.Label(row_login, text="B站登录：", font=("微软雅黑", 10), width=10, anchor="e").pack(side="left")
        self.login_status = tk.Label(row_login, text="未登录", font=("微软雅黑", 9),
                                     fg="#c0392b", width=6, anchor="w")
        self.login_status.pack(side="left", padx=4)
        tk.Button(row_login, text="🔑 打开B站登录页", command=self._open_bili_login,
                  font=("微软雅黑", 9), width=16, cursor="hand2").pack(side="left", padx=4)
        tk.Label(row_login, text="登录后粘贴 SESSDATA：", font=("微软雅黑", 9)).pack(side="left", padx=(8,2))
        self.sess_entry = tk.Entry(row_login, textvariable=self.sessdata, font=("Consolas", 9), width=30, show="*")
        self.sess_entry.pack(side="left", padx=4)
        tk.Button(row_login, text="验证", command=self._verify_bili_login,
                  font=("微软雅黑", 9), width=6).pack(side="left")
        tk.Button(row_login, text="显示/隐藏", command=self._toggle_sess_show,
                  font=("微软雅黑", 9), width=8).pack(side="left", padx=2)

        # 选项勾选
        row3 = tk.Frame(inp_frame); row3.pack(fill="x", pady=4)
        self.chk_reply = tk.BooleanVar(value=True)
        self.chk_videodm = tk.BooleanVar(value=True)
        self.chk_livedm = tk.BooleanVar(value=True)
        self.chk_medal = tk.BooleanVar(value=True)
        self.chk_titles = tk.BooleanVar(value=True)
        self.chk_ip = tk.BooleanVar(value=False)
        tk.Checkbutton(row3, text="评论", variable=self.chk_reply, font=("微软雅黑", 10)).pack(side="left", padx=6)
        tk.Checkbutton(row3, text="视频弹幕", variable=self.chk_videodm, font=("微软雅黑", 10)).pack(side="left", padx=6)
        tk.Checkbutton(row3, text="直播弹幕", variable=self.chk_livedm, font=("微软雅黑", 10)).pack(side="left", padx=6)
        tk.Checkbutton(row3, text="粉丝牌", variable=self.chk_medal, font=("微软雅黑", 10)).pack(side="left", padx=6)
        tk.Checkbutton(row3, text="查视频标题(较慢)", variable=self.chk_titles, font=("微软雅黑", 10)).pack(side="left", padx=6)
        tk.Checkbutton(row3, text="查IP属地(需登录)", variable=self.chk_ip, font=("微软雅黑", 10)).pack(side="left", padx=6)

        # ── 按钮 ──
        btn_frame = tk.Frame(self.root); btn_frame.pack(**pad)
        self.run_btn = tk.Button(btn_frame, text="▶ 开始爬取", command=self._start,
                                 font=("微软雅黑", 12, "bold"), fg="white", bg="#2B579A",
                                 width=16, height=2, cursor="hand2")
        self.run_btn.pack(side="left", padx=6)
        tk.Button(btn_frame, text="打开输出目录", command=lambda: os.startfile(self.out_dir.get()),
                  font=("微软雅黑", 10), width=14).pack(side="left", padx=6)

        # ── 进度条 ──
        self.progress = ttk.Progressbar(self.root, mode="indeterminate", length=650)
        self.progress.pack(**pad, pady=(4, 0))

        # ── 日志区 ──
        log_frame = tk.LabelFrame(self.root, text="运行日志", font=("微软雅黑", 10), padx=6, pady=4)
        log_frame.pack(fill="both", expand=True, **pad, pady=(6, 8))
        self.log_area = scrolledtext.ScrolledText(log_frame, font=("Consolas", 9), wrap="word",
                                                  bg="#1a1a2e", fg="#e0e0e0",
                                                  insertbackground="white")
        self.log_area.pack(fill="both", expand=True)

        # 底部
        info = tk.Label(self.root, text="* 动态内容 & IP属地 需B站登录态API，aicu.cc暂不支持。爬虫自动使用 curl_cffi 绕过 Cloudflare。",
                        font=("微软雅黑", 8), fg="#aaa")
        info.pack(side="bottom", pady=(0, 6))

    def _browse_dir(self):
        d = filedialog.askdirectory(title="选择输出目录", initialdir=self.out_dir.get())
        if d: self.out_dir.set(d)

    def _open_bili_login(self):
        """打开B站登录页"""
        import webbrowser
        webbrowser.open("https://passport.bilibili.com/login")
        self.log("[登录] 已在浏览器打开B站登录页，登录后按 F12 → Application → Cookies → 复制 SESSDATA 的值")

    def _verify_bili_login(self):
        """验证 B站 登录态"""
        s = self.sessdata.get().strip()
        if not s:
            messagebox.showwarning("提示", "请先粘贴 SESSDATA cookie")
            return
        try:
            resp = requests.get("https://api.bilibili.com/x/web-interface/nav",
                                headers={"User-Agent": UA, "Cookie": f"SESSDATA={s}"}, timeout=10)
            data = resp.json()
            if data.get("code") == 0 and data["data"].get("isLogin"):
                uname = data["data"]["uname"]
                self.login_status.config(text=f"✓ {uname}", fg="#27ae60")
                self.log(f"[登录] 验证成功！当前账号: {uname}")
                messagebox.showinfo("登录成功", f"已登录 B站账号: {uname}")
            else:
                self.login_status.config(text="未登录", fg="#c0392b")
                self.log("[登录] 验证失败，SESSDATA 无效或已过期")
                messagebox.showwarning("登录失败", "SESSDATA 无效或已过期，请重新获取")
        except Exception as e:
            self.log(f"[登录] 验证请求失败: {e}")

    def _toggle_sess_show(self):
        """切换 SESSDATA 显示/隐藏"""
        if self.sess_entry.cget("show") == "*":
            self.sess_entry.config(show="")
        else:
            self.sess_entry.config(show="*")

    def log(self, msg):
        self.log_area.insert("end", msg + "\n")
        self.log_area.see("end")
        self.root.update_idletasks()

    def _start(self):
        if self.running: return
        uid = self.uid_entry.get().strip()
        if not uid or not uid.isdigit():
            messagebox.showwarning("输入错误", "请输入有效的 B站 UID（纯数字）")
            return
        self.running = True
        self.run_btn.config(state="disabled", text="爬取中...")
        self.progress.start(10)
        self.log_area.delete("1.0", "end")
        os.makedirs(self.out_dir.get(), exist_ok=True)
        threading.Thread(target=self._run, args=(uid,), daemon=True).start()

    def _run(self, uid):
        out = self.out_dir.get()
        self.log(f"{'='*50}")
        self.log(f"AICU 爬虫启动 - UID: {uid}")
        self.log(f"输出目录: {out}")
        self.log(f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.log(f"{'='*50}\n")

        reply_raw = []; videodm_raw = []; livedm_flat = []; medal_rows = []
        all_oids = set()

        try:
            # 1) 评论
            if self.chk_reply.get():
                reply_raw = crawl_reply(uid, self.log)
                if reply_raw:
                    for c in reply_raw:
                        o = c.get("dyn",{}).get("oid","")
                        if o: all_oids.add(str(o))
                time.sleep(REQUEST_DELAY)

            # 2) 视频弹幕
            if self.chk_videodm.get():
                videodm_raw = crawl_videodm(uid, self.log)
                if videodm_raw:
                    for dm in videodm_raw:
                        o = dm.get("oid","")
                        if o: all_oids.add(str(o))
                time.sleep(REQUEST_DELAY)

            # 3) 直播弹幕
            if self.chk_livedm.get():
                livedm_flat = crawl_livedm(uid, self.log)
                time.sleep(REQUEST_DELAY)

            # 4) 粉丝牌
            if self.chk_medal.get():
                medal_rows = crawl_medal(uid, self.log)

            # ── 视频标题 ──
            titles = {}
            if self.chk_titles.get() and all_oids:
                self.log(f"\n[视频标题] 去重后 {len(all_oids)} 个视频...")
                titles = fetch_titles(list(all_oids), out, self.log)
                self.log("[视频标题] 完成")

            # ── IP属地解析 ──
            ip_map = {}
            if reply_raw and self.chk_ip.get() and self.sessdata.get().strip():
                self.log("\n[IP属地] 检测到已登录，正在解析IP属地...")
                ip_map = resolve_ips(reply_raw, self.sessdata.get().strip(), self.log)

            # ── 整理 ──
            reply_rows = flatten_reply(reply_raw, titles, ip_map) if reply_raw else []
            videodm_rows = flatten_videodm(videodm_raw, titles) if videodm_raw else []

            # ── 生成 Excel ──
            self.log("\n[Excel] 正在生成...")
            path = build_excel(uid, reply_rows, videodm_rows, livedm_flat, medal_rows, out)
            self.log(f"[Excel] 已保存: {path}\n")

            # 摘要
            self.log("=" * 50)
            self.log("完成摘要:")
            if reply_rows: self.log(f"  评论:     {len(reply_rows)} 条")
            if videodm_rows: self.log(f"  视频弹幕: {len(videodm_rows)} 条")
            if livedm_flat: self.log(f"  直播弹幕: {len(livedm_flat)} 条")
            if medal_rows: self.log(f"  粉丝牌:   {len(medal_rows)} 个")
            if titles: self.log(f"  视频标题: {len(titles)} 个")
            self.log("=" * 50)
            self.log("全部完成！")

        except Exception as e:
            self.log(f"\n[ERROR] {e}")
            traceback.print_exc(file=sys.stderr)
        finally:
            self.running = False
            self.root.after(0, self._done)

    def _done(self):
        self.run_btn.config(state="normal", text="▶ 开始爬取")
        self.progress.stop()

    def run(self):
        self.root.mainloop()


# ═══════════════════════════════════════════════════
if __name__ == "__main__":
    AicuApp().run()
